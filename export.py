"""Excel / JSON export utilities."""
import os
import json
import shutil
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from db import db_query

EXPORT_DIR = os.path.join(os.path.dirname(__file__), "exports")
LATEST_DIR_NAME = "_latest"
LEGACY_DIR_NAME = "_历史散文件"
EXPORT_FOLDERS = {
    "products": ("产品列表", "产品列表_最新.xlsx"),
    "recipes": ("配方记录", "配方记录_最新.xlsx"),
    "success_rate": ("成功率", "成功率_最新.xlsx"),
    "json": ("JSON数据", "JSON数据_最新.json"),
    "templates": ("导入模板", ""),
}
LEGACY_PREFIXES = ("products_", "recipes_", "success_rate_", "export_")


def _ensure_dir():
    os.makedirs(EXPORT_DIR, exist_ok=True)
    os.makedirs(os.path.join(EXPORT_DIR, LATEST_DIR_NAME), exist_ok=True)
    _move_legacy_flat_exports()


def _move_legacy_flat_exports():
    legacy_dir = os.path.join(EXPORT_DIR, LEGACY_DIR_NAME)
    for name in os.listdir(EXPORT_DIR):
        path = os.path.join(EXPORT_DIR, name)
        if not os.path.isfile(path):
            continue
        if not name.startswith(LEGACY_PREFIXES):
            continue
        os.makedirs(legacy_dir, exist_ok=True)
        shutil.move(path, os.path.join(legacy_dir, name))


def _export_path(export_key: str, filename: str) -> str:
    folder_name, _ = EXPORT_FOLDERS[export_key]
    folder = os.path.join(EXPORT_DIR, folder_name)
    os.makedirs(folder, exist_ok=True)
    return os.path.join(folder, filename)


def _copy_latest(path: str, export_key: str) -> None:
    _, latest_name = EXPORT_FOLDERS[export_key]
    latest_path = os.path.join(EXPORT_DIR, LATEST_DIR_NAME, latest_name)
    shutil.copy2(path, latest_path)


def _now_str() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _set_header(ws, headers):
    thin = Side(style="thin", color="D8DEE4")
    header_fill = PatternFill(start_color="F6F8FA", end_color="F6F8FA", fill_type="solid")
    for col, val in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=val)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = Border(bottom=thin)
        cell.alignment = Alignment(horizontal="left", vertical="center")


def export_excel(export_type: str) -> str:
    _ensure_dir()
    wb = Workbook()
    ws = wb.active
    now = _now_str()

    if export_type == "products":
        ws.title = "产品列表"
        _set_header(ws, ["品号", "规格", "品名", "当前数量", "状态", "创建时间"])
        rows = db_query("SELECT 品号, 规格, 品名, 当前数量, 状态, created_at FROM products ORDER BY created_at DESC")
        for ridx, row in enumerate(rows, 2):
            for cidx, key in enumerate(["品号", "规格", "品名", "当前数量", "状态", "created_at"], 1):
                ws.cell(row=ridx, column=cidx, value=row[key])
        filename = f"products_{now}.xlsx"
        export_key = "products"

    elif export_type == "recipes":
        ws.title = "配方记录"
        _set_header(ws, ["产品id", "产品品号", "产品品名", "试验日期", "配方名称", "状态", "用了多少", "备注", "创建时间"])
        rows = db_query("""
            SELECT r.产品id, p.品号, p.品名, r.试验日期, r.配方名称, r.状态, r.用了多少, r.备注, r.created_at
            FROM recipes r
            JOIN products p ON r.产品id = p.id
            ORDER BY r.试验日期 DESC
        """)
        for ridx, row in enumerate(rows, 2):
            for cidx, key in enumerate(["产品id", "品号", "品名", "试验日期", "配方名称", "状态", "用了多少", "备注", "created_at"], 1):
                ws.cell(row=ridx, column=cidx, value=row[key])
        filename = f"recipes_{now}.xlsx"
        export_key = "recipes"

    elif export_type == "success_rate":
        ws.title = "成功率汇总"
        _set_header(ws, ["产品品号", "产品品名", "配方hash", "试验次数", "成功次数", "成功率"])
        rows = db_query("""
            SELECT v.*, p.品号, p.品名
            FROM v_recipe_success_rate v
            JOIN products p ON v.产品id = p.id
            ORDER BY v.成功率 DESC, v.试验次数 DESC
        """)
        for ridx, row in enumerate(rows, 2):
            for cidx, key in enumerate(["品号", "品名", "配方hash", "试验次数", "成功次数", "成功率"], 1):
                ws.cell(row=ridx, column=cidx, value=row[key])
        filename = f"success_rate_{now}.xlsx"
        export_key = "success_rate"
    else:
        raise ValueError(f"Unknown export type: {export_type}")

    path = _export_path(export_key, filename)
    wb.save(path)
    _copy_latest(path, export_key)
    return path


def export_import_template(template_type: str) -> str:
    _ensure_dir()
    wb = Workbook()
    ws = wb.active
    export_key = "templates"

    if template_type == "products":
        ws.title = "产品导入模板"
        headers = ["品号", "规格", "品名", "当前数量", "备注"]
        sample = ["P-001", "100g", "样品名称", 10, "可选"]
        filename = "products_import_template.xlsx"
    elif template_type == "recipes":
        ws.title = "配方导入模板"
        headers = ["产品品号", "试验日期", "配方名称", "状态", "用了多少", "类型", "名称", "用量", "单位", "备注"]
        sample = ["P-001", "2026-06-07", "配方A", "pending", 1, "原料", "水", 1, "g", "状态可填 success/failed/pending"]
        filename = "recipes_import_template.xlsx"
    else:
        raise ValueError(f"Unknown template type: {template_type}")

    _set_header(ws, headers)
    for col, value in enumerate(sample, 1):
        ws.cell(row=2, column=col, value=value)
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 18

    path = _export_path(export_key, filename)
    wb.save(path)
    return path


def export_json() -> str:
    _ensure_dir()
    now = _now_str()
    data = {
        "exported_at": datetime.now().isoformat(),
        "products": [dict(r) for r in db_query("SELECT * FROM products ORDER BY id")],
        "recipes": [dict(r) for r in db_query("SELECT * FROM recipes ORDER BY id")],
        "materials": [dict(r) for r in db_query("SELECT * FROM recipe_materials ORDER BY id")],
    }
    filename = f"export_{now}.json"
    path = _export_path("json", filename)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    _copy_latest(path, "json")
    return path
