"""Checks for DESIGN-v2 feature completion gaps."""
import os
import re
import sqlite3
import sys
import tempfile

from openpyxl import load_workbook
from fastapi.testclient import TestClient

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

import app as app_module
import db as db_module
import export as export_module


def use_temp_db():
    tmpdir = tempfile.TemporaryDirectory()
    db_path = os.path.join(tmpdir.name, "kitchen.db")
    db_module.DB_PATH = db_path
    app_module.DB_PATH = db_path
    db_module.init_db()
    return tmpdir


def create_product(client, code, name="测试产品"):
    response = client.post(
        "/api/products",
        json={"品号": code, "规格": "100g", "品名": name, "当前数量": 1},
        headers={"X-Username": "tester"},
    )
    assert response.status_code == 200, response.text
    return response.json()["data"]["id"]


def create_recipe(client, product_id, date, status, material_name="水", recipe_name="配方"):
    response = client.post(
        "/api/recipes",
        json={
            "产品id": product_id,
            "试验日期": date,
            "配方名称": recipe_name,
            "状态": status,
            "用了多少": 1,
            "备注": f"{date}-{status}",
            "原料辅料": [
                {"类型": "原料", "名称": material_name, "用量": 1, "单位": "g"}
            ],
        },
        headers={"X-Username": "tester"},
    )
    assert response.status_code == 200, response.text
    return response.json()["data"]["id"]


def extract_js_method_block(js, method_name, next_method_name):
    start_marker = f"  {method_name}"
    end_marker = f"\n  {next_method_name}"
    start = js.find(start_marker)
    assert start != -1, f"missing {method_name} block"
    end = js.find(end_marker, start)
    assert end != -1, f"missing {method_name} end boundary"
    return js[start:end]


def check_success_rate_filters_and_sort(client):
    product_id = create_product(client, "P-SR-A", "成功率产品A")
    other_product_id = create_product(client, "P-SR-B", "成功率产品B")
    create_recipe(client, product_id, "2026-06-01", "success", "水", "A-1")
    create_recipe(client, product_id, "2026-06-02", "failed", "水", "A-2")
    create_recipe(client, product_id, "2026-05-01", "success", "水", "A-old")
    create_recipe(client, other_product_id, "2026-06-03", "success", "糖", "B-1")

    response = client.get(
        "/api/success-rate",
        params={
            "product_id": product_id,
            "date_from": "2026-06-01",
            "date_to": "2026-06-30",
            "min_rate": "0.5",
            "max_rate": "0.5",
            "sort": "成功率",
            "order": "asc",
        },
    )
    assert response.status_code == 200, response.text
    items = response.json()["data"]
    assert len(items) == 1, items
    assert items[0]["产品id"] == product_id
    assert items[0]["试验次数"] == 2
    assert items[0]["成功次数"] == 1
    assert items[0]["成功率"] == 0.5

    only_success = client.get(
        "/api/success-rate",
        params={"product_id": product_id, "status": "success"},
    )
    assert only_success.status_code == 200, only_success.text
    success_items = only_success.json()["data"]
    assert success_items[0]["试验次数"] == 2
    assert success_items[0]["成功次数"] == 2
    assert success_items[0]["成功率"] == 1.0


def check_recipe_success_rate_range(client):
    product_id = create_product(client, "P-RANGE", "配方范围产品")
    kept_id = create_recipe(client, product_id, "2026-06-01", "success", "蜂蜜", "高成功")
    create_recipe(client, product_id, "2026-06-02", "success", "蜂蜜", "高成功-2")
    create_recipe(client, product_id, "2026-06-03", "failed", "盐", "低成功")

    response = client.get(
        "/api/recipes",
        params={
            "product_id": product_id,
            "min_rate": "1",
            "max_rate": "1",
            "sort": "成功率",
            "order": "desc",
        },
    )
    assert response.status_code == 200, response.text
    items = response.json()["data"]["items"]
    assert {item["id"] for item in items} == {kept_id, kept_id + 1}
    assert all(item["成功率"] == 1.0 for item in items)


def check_same_hash_history(client):
    product_id = create_product(client, "P-HASH", "历史产品")
    older_id = create_recipe(client, product_id, "2026-06-01", "success", "柠檬", "第一次")
    current_id = create_recipe(client, product_id, "2026-06-03", "failed", "柠檬", "第二次")
    create_recipe(client, product_id, "2026-06-02", "success", "薄荷", "不同组合")

    response = client.get(f"/api/recipes/{current_id}/same-hash")
    assert response.status_code == 200, response.text
    items = response.json()["data"]
    assert [item["id"] for item in items] == [current_id, older_id]
    assert [item["状态"] for item in items] == ["failed", "success"]

    detail = client.get(f"/api/recipes/{current_id}").json()["data"]
    by_hash = client.get(f"/api/products/{product_id}/recipes-by-hash/{detail['配方hash']}")
    assert by_hash.status_code == 200, by_hash.text
    by_hash_items = by_hash.json()["data"]
    assert [item["id"] for item in by_hash_items] == [current_id, older_id]


def check_recipe_excel_includes_product_id(client, tmpdir):
    product_id = create_product(client, "P-XLSX", "导出产品")
    create_recipe(client, product_id, "2026-06-01", "success", "水", "导出配方")
    export_module.EXPORT_DIR = tmpdir.name

    path = export_module.export_excel("recipes")
    wb = load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    assert headers[:2] == ["产品id", "产品品号"], headers
    rows = [[cell.value for cell in row] for row in ws.iter_rows(min_row=2)]
    matching = [row for row in rows if row[1] == "P-XLSX"]
    assert len(matching) == 1, rows
    assert matching[0][0] == product_id


def check_frontend_design_contracts():
    with open(os.path.join(os.path.dirname(os.path.dirname(__file__)), "static", "app.js"), encoding="utf-8") as f:
        js = f.read()
    with open(os.path.join(os.path.dirname(os.path.dirname(__file__)), "static", "style.css"), encoding="utf-8") as f:
        css = f.read()
    with open(os.path.join(os.path.dirname(os.path.dirname(__file__)), "templates", "base.html"), encoding="utf-8") as f:
        base_html = f.read()
    frontend_bundle = "\n".join([js, css, base_html])

    required_snippets = [
        "routeFromLocation",
        "location.pathname",
        "renderPageHero",
        "renderStatusBar",
        "renderTabs",
        "products-hero-search",
        "openInventoryModal",
        "submitInventoryModal",
        "inventory-modal",
        "inventory-delta",
        "inventory-reason",
        "inventory-error",
        "recipe-editor-shell",
        "recipe-section",
        "success-rate-hero",
        "renderEmptyState",
        "prod-rec-status",
        "prod-rec-date-from",
        "prod-rec-date-to",
        "prod-rec-min-rate",
        "prod-rec-sort",
        "sr-date-from",
        "sr-date-to",
        "sr-status",
        "min_rate",
        "max_rate",
        "same-hash",
        "recipes-by-hash",
        "toggleHashHistory",
        "确认备份",
        "app.archiveProduct",
        "app.restoreProduct",
        "app.deleteProduct",
        "modal-body",
        "toast-title",
        "toast-message",
        "app-shell",
        "navbar-expand-lg",
        "page-wrapper",
        "page-header",
        "workbench-tabs",
        "workbench-body",
        "statusbar",
        "export-menu",
        "sidebar-brand",
        "sidebar-subnav",
        "topbar-actions",
        "workbench-tab",
        "renderExportMenu",
        "renderSidebarNav",
        "renderTopbarActions",
        "renderStatusBar",
        "copyText",
        "renderLocalIcons(root = document)",
        "renderIcon(name, className = '')",
        "renderBrandMark()",
        "sidebar-brand-logo",
        "brand-fallback",
        "/兴达logo.ico",
    ]
    for snippet in required_snippets:
        assert snippet in frontend_bundle, f"missing frontend contract: {snippet}"
    recipe_detail_body = extract_js_method_block(js, "async renderRecipeDetail(el, id)", "deleteRecipe(")
    hero_match = re.search(r"const hero = this\.renderPageHero\(\{(?P<body>[\s\S]*?)\}\);", recipe_detail_body)
    assert hero_match, "missing renderRecipeDetail hero block"
    hero_body = hero_match.group("body")
    assert "escapeHtml" not in hero_body, "recipe detail hero should rely on renderPageHero escaping"
    assert "app.copyText(" in recipe_detail_body, "recipe detail should copy hash via helper"
    assert "navigator.clipboard.writeText" not in recipe_detail_body, "recipe detail should not inline clipboard calls"

    history_panel_match = re.search(
        r"同组合历史[\s\S]*?\$\{(?P<expr>historyItems\.length\s*\?[\s\S]*?)\}",
        recipe_detail_body,
    )
    assert history_panel_match, "history panel should be controlled by historyItems.length"

    toast_body = extract_js_method_block(js, "toast(msg, type='success')", "async copyText(")
    assert "toast-title" in toast_body and "toast-message" in toast_body

    topbar_body = extract_js_method_block(js, "renderTopbarActions()", "renderPageHero(")
    assert "renderExportMenu()" in topbar_body
    assert "renderIcon(" in topbar_body

    sidebar_body = extract_js_method_block(js, "renderSidebarNav(path)", "sidebarSearch(")
    assert "renderBrandMark()" in sidebar_body
    assert "renderIcon(" in sidebar_body

    brand_body = extract_js_method_block(js, "renderBrandMark()", "renderStatusBar(")
    assert "sidebar-brand-logo" in brand_body
    assert "brand-fallback" in brand_body
    assert "navbar-expand-lg" in base_html, "base shell should keep sidebar drawer scaffold"
    assert 'id="app-shell"' in base_html and 'id="sidebar"' in base_html
    assert 'id="workbench-tabs"' in base_html and 'id="workbench-body"' in base_html


def main():
    tmpdir = use_temp_db()
    try:
        client = TestClient(app_module.app)
        check_success_rate_filters_and_sort(client)
        check_recipe_success_rate_range(client)
        check_same_hash_history(client)
        check_recipe_excel_includes_product_id(client, tmpdir)
        check_frontend_design_contracts()
    finally:
        tmpdir.cleanup()


if __name__ == "__main__":
    main()
