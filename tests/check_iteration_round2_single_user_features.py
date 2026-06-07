"""Checks for single-user productivity features from the second iteration."""
import os
import sys
import tempfile

from fastapi.testclient import TestClient
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

import app as app_module
import backup as backup_module
import db as db_module
import export as export_module


def use_temp_paths():
    tmpdir = tempfile.TemporaryDirectory()
    db_path = os.path.join(tmpdir.name, "kitchen.db")
    db_module.DB_PATH = db_path
    backup_module.DB_PATH = db_path
    export_module.EXPORT_DIR = os.path.join(tmpdir.name, "exports")
    db_module.init_db()
    return tmpdir, export_module.EXPORT_DIR


def create_product(client, code, name="二轮产品"):
    response = client.post(
        "/api/products",
        json={"品号": code, "规格": "100g", "品名": name, "当前数量": 10},
        headers={"X-Username": "tester"},
    )
    assert response.status_code == 200, response.text
    return response.json()["data"]["id"]


def create_recipe(client, product_id, date, status, material, name="配方"):
    response = client.post(
        "/api/recipes",
        json={
            "产品id": product_id,
            "试验日期": date,
            "配方名称": name,
            "状态": status,
            "用了多少": 1,
            "原料辅料": [
                {"类型": "原料", "名称": material, "用量": 1, "单位": "g"},
                {"类型": "辅料", "名称": "盐", "用量": 0.1, "单位": "g"},
            ],
        },
        headers={"X-Username": "tester"},
    )
    assert response.status_code == 200, response.text
    return response.json()["data"]["id"]


def check_recipe_duplicate_preserves_materials(client):
    product_id = create_product(client, "P-COPY")
    source_id = create_recipe(client, product_id, "2026-06-01", "success", "水", "原始配方")

    response = client.post(
        f"/api/recipes/{source_id}/duplicate",
        json={"试验日期": "2026-06-08", "配方名称": "复测配方"},
        headers={"X-Username": "tester"},
    )
    assert response.status_code == 200, response.text
    new_id = response.json()["data"]["id"]

    source = client.get(f"/api/recipes/{source_id}").json()["data"]
    copied = client.get(f"/api/recipes/{new_id}").json()["data"]
    assert copied["产品id"] == product_id
    assert copied["试验日期"] == "2026-06-08"
    assert copied["配方名称"] == "复测配方"
    assert copied["状态"] == "pending"
    assert copied["配方hash"] == source["配方hash"]
    assert [(m["类型"], m["名称"], m["用量"], m["单位"]) for m in copied["原料辅料"]] == [
        (m["类型"], m["名称"], m["用量"], m["单位"]) for m in source["原料辅料"]
    ]


def check_material_suggestions_and_reverse_lookup(client):
    product_id = create_product(client, "P-MATERIAL-A")
    other_product_id = create_product(client, "P-MATERIAL-B")
    water_id = create_recipe(client, product_id, "2026-06-01", "success", "水", "含水配方")
    create_recipe(client, product_id, "2026-06-02", "failed", "糖", "含糖配方")
    create_recipe(client, other_product_id, "2026-06-03", "success", "水", "另一含水配方")

    suggestions = client.get("/api/materials/suggestions")
    assert suggestions.status_code == 200, suggestions.text
    items = suggestions.json()["data"]
    water = next(item for item in items if item["名称"] == "水")
    assert water["使用次数"] == 2
    assert water["单位"] == "g"

    by_material = client.get("/api/recipes", params={"material": "水", "sort": "试验日期"})
    assert by_material.status_code == 200, by_material.text
    recipe_ids = {item["id"] for item in by_material.json()["data"]["items"]}
    assert water_id in recipe_ids
    assert len(recipe_ids) == 2


def check_success_rate_min_trials_filter(client):
    product_id = create_product(client, "P-MIN-TRIALS")
    create_recipe(client, product_id, "2026-06-01", "success", "蜂蜜", "蜂蜜一")
    create_recipe(client, product_id, "2026-06-02", "failed", "蜂蜜", "蜂蜜二")
    create_recipe(client, product_id, "2026-06-03", "success", "柠檬", "柠檬一")

    response = client.get("/api/success-rate", params={"product_id": product_id, "min_trials": 2})
    assert response.status_code == 200, response.text
    items = response.json()["data"]
    assert len(items) == 1, items
    assert items[0]["试验次数"] == 2


def check_inventory_adjustment_records_movement(client):
    product_id = create_product(client, "P-STOCK")
    response = client.post(
        f"/api/products/{product_id}/inventory-adjust",
        json={"变动数量": -3, "原因": "试验消耗"},
        headers={"X-Username": "tester"},
    )
    assert response.status_code == 200, response.text
    data = response.json()["data"]
    assert data["变动前"] == 10
    assert data["变动后"] == 7

    product = client.get(f"/api/products/{product_id}").json()["data"]
    assert product["当前数量"] == 7

    movements = client.get(f"/api/products/{product_id}/inventory-movements")
    assert movements.status_code == 200, movements.text
    items = movements.json()["data"]
    assert len(items) == 1
    assert items[0]["变动数量"] == -3
    assert items[0]["原因"] == "试验消耗"


def check_import_templates_have_expected_headers(client, export_dir):
    for template_type, expected_headers in {
        "products": ["品号", "规格", "品名", "当前数量", "备注"],
        "recipes": ["产品品号", "试验日期", "配方名称", "状态", "用了多少", "类型", "名称", "用量", "单位", "备注"],
    }.items():
        response = client.get("/api/export/template", params={"type": template_type})
        assert response.status_code == 200, response.text
        path = os.path.join(export_dir, "导入模板", f"{template_type}_import_template.xlsx")
        assert os.path.exists(path), path
        wb = load_workbook(path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert headers == expected_headers, headers


def check_frontend_exposes_second_round_controls():
    root = os.path.dirname(os.path.dirname(__file__))
    with open(os.path.join(root, "static", "app.js"), encoding="utf-8") as f:
        js = f.read()
    with open(os.path.join(root, "templates", "base.html"), encoding="utf-8") as f:
        html = f.read()

    for snippet in [
        "duplicateRecipe",
        "inventory-adjust",
        "inventory-movements",
        "exportTemplate",
        "导入模板",
        "prod-rec-material",
        "sr-min-trials",
    ]:
        assert snippet in js or snippet in html, f"missing frontend control: {snippet}"


def main():
    checks = [
        check_recipe_duplicate_preserves_materials,
        check_material_suggestions_and_reverse_lookup,
        check_success_rate_min_trials_filter,
        check_inventory_adjustment_records_movement,
    ]
    for check in checks:
        tmpdir, _ = use_temp_paths()
        try:
            client = TestClient(app_module.app)
            check(client)
        finally:
            tmpdir.cleanup()
    tmpdir, export_dir = use_temp_paths()
    try:
        client = TestClient(app_module.app)
        check_import_templates_have_expected_headers(client, export_dir)
    finally:
        tmpdir.cleanup()
    check_frontend_exposes_second_round_controls()


if __name__ == "__main__":
    main()
